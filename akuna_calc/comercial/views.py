from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Q
from django.http import JsonResponse
from datetime import datetime
from decimal import Decimal
from .models import Cliente, Venta, Cuenta, Compra, TipoCuenta, TipoGasto, PagoVenta, PagoCompra
from .forms import ClienteForm, VentaForm, CuentaForm, CompraForm, ReporteForm, ReporteProveedorForm


def _resolver_monto_pago_venta(monto, pago_en_dolares, monto_usd, cotizacion_usd):
    monto_usd_decimal = Decimal(monto_usd) if monto_usd else None
    cotizacion_usd_decimal = Decimal(cotizacion_usd) if cotizacion_usd else None

    if monto not in (None, ''):
        monto_decimal = Decimal(monto)
    elif pago_en_dolares and monto_usd_decimal and cotizacion_usd_decimal:
        monto_decimal = (monto_usd_decimal * cotizacion_usd_decimal).quantize(Decimal('0.01'))
    else:
        raise ValueError('Debe indicar el monto en pesos o completar monto USD y cotizacion.')

    return monto_decimal, monto_usd_decimal, cotizacion_usd_decimal


@login_required
def dashboard_comercial(request):
    from datetime import datetime, timedelta
    from django.db.models.functions import TruncMonth
    import json
    from django.core.serializers.json import DjangoJSONEncoder
    
    # Estad�sticas generales
    total_ventas = Venta.objects.filter(deleted_at__isnull=True).aggregate(Sum('valor_total'))['valor_total__sum'] or 0
    total_compras = Compra.objects.filter(deleted_at__isnull=True).aggregate(Sum('valor_total'))['valor_total__sum'] or 0
    ventas_pendientes = Venta.objects.filter(deleted_at__isnull=True, estado='pendiente').count()
    
    # �ltimos 6 meses
    hace_6_meses = datetime.now() - timedelta(days=180)
    
    # Ventas por mes
    ventas_por_mes = Venta.objects.filter(deleted_at__isnull=True, created_at__gte=hace_6_meses).annotate(
        mes=TruncMonth('created_at')
    ).values('mes').annotate(
        total=Sum('valor_total')
    ).order_by('mes')
    
    # Compras por mes
    compras_por_mes = Compra.objects.filter(deleted_at__isnull=True, fecha_pago__gte=hace_6_meses).annotate(
        mes=TruncMonth('fecha_pago')
    ).values('mes').annotate(
        total=Sum('valor_total')
    ).order_by('mes')
    
    # Top 5 clientes
    top_clientes = Venta.objects.filter(deleted_at__isnull=True).values('cliente__nombre', 'cliente__apellido').annotate(
        total=Sum('valor_total')
    ).order_by('-total')[:5]
    
    # Compras por tipo de cuenta
    compras_por_tipo_raw = Compra.objects.filter(deleted_at__isnull=True).values('cuenta__tipo_cuenta__id', 'cuenta__tipo_cuenta__tipo').annotate(
        total=Sum('valor_total')
    ).order_by('-total')
    
    # Formatear datos para el gr�fico
    compras_por_tipo = []
    for item in compras_por_tipo_raw:
        tipo_cuenta = TipoCuenta.objects.filter(id=item['cuenta__tipo_cuenta__id']).first()
        if tipo_cuenta:
            compras_por_tipo.append({
                'tipo': tipo_cuenta.get_tipo_display(),
                'total': float(item['total'])
            })
    
    context = {
        'total_ventas': total_ventas,
        'total_compras': total_compras,
        'ventas_pendientes': ventas_pendientes,
        'clientes_count': Cliente.objects.filter(deleted_at__isnull=True).count(),
        'ventas_por_mes': json.dumps(list(ventas_por_mes), cls=DjangoJSONEncoder),
        'compras_por_mes': json.dumps(list(compras_por_mes), cls=DjangoJSONEncoder),
        'top_clientes': list(top_clientes),
        'compras_por_tipo': json.dumps(compras_por_tipo),
    }
    return render(request, 'comercial/dashboard.html', context)


def construir_cuenta_corriente_proveedor(proveedor):
    compras = list(
        Compra.objects.filter(
            deleted_at__isnull=True,
            cuenta=proveedor,
        ).prefetch_related('pagos_compra').order_by('fecha_pago', 'created_at', 'pk')
    )

    movimientos = []
    saldo_acumulado = Decimal('0')

    for compra in compras:
        saldo_acumulado += compra.valor_total
        movimientos.append({
            'fecha': compra.fecha_pago,
            'tipo': 'compra',
            'descripcion': f'Compra #{compra.numero_pedido}',
            'referencia': compra.numero_factura or compra.numero_pedido,
            'forma_pago': '-',
            'debe': compra.valor_total,
            'haber': Decimal('0'),
            'saldo': saldo_acumulado,
            'compra': compra,
        })

        if compra.sena > 0:
            saldo_acumulado -= compra.sena
            movimientos.append({
                'fecha': compra.fecha_pago,
                'tipo': 'sena',
                'descripcion': f'Seña inicial de compra #{compra.numero_pedido}',
                'referencia': compra.numero_factura or compra.numero_pedido,
                'forma_pago': compra.get_forma_pago_sena_display() if compra.forma_pago_sena else 'Sin informar',
                'debe': Decimal('0'),
                'haber': compra.sena,
                'saldo': saldo_acumulado,
                'compra': compra,
            })

        for pago in compra.pagos_compra.all().order_by('fecha_pago', 'created_at', 'pk'):
            saldo_acumulado -= pago.monto
            movimientos.append({
                'fecha': pago.fecha_pago,
                'tipo': 'pago',
                'descripcion': f'Pago de compra #{compra.numero_pedido}',
                'referencia': pago.numero_factura or compra.numero_pedido,
                'forma_pago': pago.get_forma_pago_display(),
                'debe': Decimal('0'),
                'haber': pago.monto,
                'saldo': saldo_acumulado,
                'compra': compra,
                'pago': pago,
            })

    movimientos.sort(key=lambda item: (item['fecha'], 0 if item['tipo'] == 'compra' else 1, item['descripcion']))

    saldo_acumulado = Decimal('0')
    for movimiento in movimientos:
        saldo_acumulado += movimiento['debe'] - movimiento['haber']
        movimiento['saldo'] = saldo_acumulado

    total_compras = sum(compra.valor_total for compra in compras)
    total_senas = sum(compra.sena for compra in compras)
    total_pagos = sum(sum(pago.monto for pago in compra.pagos_compra.all()) for compra in compras)
    saldo_actual = total_compras - total_senas - total_pagos

    return {
        'proveedor': proveedor,
        'compras': compras,
        'movimientos': movimientos,
        'total_compras': total_compras,
        'total_senas': total_senas,
        'total_pagos': total_pagos,
        'saldo_actual': saldo_actual,
    }


# VENTAS
@login_required
def ventas_list(request):
    from django.core.paginator import Paginator
    from django.db.models import Case, When, Value, IntegerField, Sum, Count

    ventas = Venta.objects.filter(deleted_at__isnull=True).select_related('cliente').prefetch_related('pagos', 'percepciones').all()

    # Filtros
    estado = request.GET.get('estado', '')
    con_factura = request.GET.get('con_factura', '')
    buscar = request.GET.get('q', '')
    razon_social = request.GET.get('razon_social', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    tipo_fecha = request.GET.get('tipo_fecha', 'pago')
    orden = request.GET.get('orden', '-created_at')
    con_saldo = request.GET.get('con_saldo', '')

    if estado:
        ventas = ventas.filter(estado=estado)

    if con_factura == 'si':
        ventas = ventas.filter(con_factura=True)
    elif con_factura == 'no':
        ventas = ventas.filter(con_factura=False)

    if buscar:
        ventas = ventas.filter(
            Q(numero_pedido__icontains=buscar) |
            Q(cliente__nombre__icontains=buscar) |
            Q(cliente__apellido__icontains=buscar) |
            Q(cliente__razon_social__icontains=buscar) |
            Q(numero_factura__icontains=buscar) |
            Q(pagos__numero_factura__icontains=buscar)
        ).distinct()

    if razon_social:
        ventas = ventas.filter(cliente__razon_social__icontains=razon_social)

    if fecha_desde:
        if tipo_fecha == 'factura':
            ventas = ventas.filter(fecha_factura__gte=fecha_desde)
        else:
            ventas = ventas.filter(fecha_pago__gte=fecha_desde)

    if fecha_hasta:
        if tipo_fecha == 'factura':
            ventas = ventas.filter(fecha_factura__lte=fecha_hasta)
        else:
            ventas = ventas.filter(fecha_pago__lte=fecha_hasta)

    if con_saldo == 'si':
        ventas = ventas.filter(saldo__gt=0)

    # Totales del filtro activo (antes de paginar)
    totales = ventas.aggregate(
        total_monto=Sum('valor_total'),
        total_saldo=Sum('saldo'),
        total_count=Count('id'),
    )
    total_monto = totales['total_monto'] or 0
    total_saldo = totales['total_saldo'] or 0
    total_count = totales['total_count'] or 0

    # Ordenamiento especial para numero_factura (manejar valores vac�os)
    if orden in ['numero_factura', '-numero_factura']:
        ventas = ventas.annotate(
            factura_order=Case(
                When(numero_factura='', then=Value(1)),
                default=Value(0),
                output_field=IntegerField()
            )
        )
        if orden == 'numero_factura':
            ventas = ventas.order_by('factura_order', 'numero_factura')
        else:
            ventas = ventas.order_by('factura_order', '-numero_factura')
    else:
        ventas = ventas.order_by(orden)

    # Paginaci�n
    paginator = Paginator(ventas, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'ventas': page_obj,
        'page_obj': page_obj,
        'estados': Venta.ESTADO_CHOICES,
        'filtro_estado': estado,
        'filtro_factura': con_factura,
        'filtro_con_saldo': con_saldo,
        'buscar': buscar,
        'razon_social': razon_social,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'tipo_fecha': tipo_fecha,
        'orden_actual': orden,
        'razones_sociales': Cliente.objects.filter(deleted_at__isnull=True, razon_social__isnull=False).exclude(razon_social='').values_list('razon_social', flat=True).distinct().order_by('razon_social'),
        'total_monto': total_monto,
        'total_saldo': total_saldo,
        'total_count': total_count,
        'col_estados': [
            ('pendiente', 'bg-yellow-400', 'bg-yellow-50', 'border-yellow-200', 'bg-yellow-100 text-yellow-800', 'bg-gradient-to-r from-amber-400 to-yellow-500'),
            ('entregado', 'bg-blue-400',   'bg-blue-50',   'border-blue-200',   'bg-blue-100 text-blue-800',   'bg-gradient-to-r from-blue-500 to-blue-600'),
            ('colocado',  'bg-green-400',  'bg-green-50',  'border-green-200',  'bg-green-100 text-green-800', 'bg-gradient-to-r from-emerald-500 to-green-600'),
        ],
    }
    return render(request, 'comercial/ventas/list.html', context)


@login_required
def venta_create(request):
    if request.method == 'POST':
        form = VentaForm(request.POST)
        if form.is_valid():
            # Validar factura duplicada
            numero_factura = form.cleaned_data.get('numero_factura')
            tipo_factura = form.cleaned_data.get('tipo_factura')
            
            if numero_factura:
                venta_existente = Venta.objects.filter(
                    numero_factura=numero_factura,
                    tipo_factura=tipo_factura,
                    deleted_at__isnull=True
                ).exists()
                
                if venta_existente:
                    messages.error(request, f'No se puede cargar esta factura. La factura {numero_factura} tipo {tipo_factura} ya se encuentra cargada en el sistema.')
                    return render(request, 'comercial/ventas/form.html', {'form': form, 'title': 'Nueva Venta'})
            
            form.save()
            messages.success(request, 'Venta creada exitosamente.')
            return redirect('comercial:ventas_list')
    else:
        form = VentaForm()
    return render(request, 'comercial/ventas/form.html', {'form': form, 'title': 'Nueva Venta'})


@login_required
def venta_edit(request, pk):
    venta = get_object_or_404(Venta, pk=pk)
    if request.method == 'POST':
        form = VentaForm(request.POST, instance=venta)
        if form.is_valid():
            # Validar factura duplicada (excluyendo la venta actual)
            numero_factura = form.cleaned_data.get('numero_factura')
            tipo_factura = form.cleaned_data.get('tipo_factura')
            
            if numero_factura:
                venta_existente = Venta.objects.filter(
                    numero_factura=numero_factura,
                    tipo_factura=tipo_factura,
                    deleted_at__isnull=True
                ).exclude(pk=pk).exists()
                
                if venta_existente:
                    messages.error(request, f'No se puede actualizar con esta factura. La factura {numero_factura} tipo {tipo_factura} ya se encuentra cargada en el sistema.')
                    return render(request, 'comercial/ventas/form.html', {'form': form, 'title': 'Editar Venta'})
            
            form.save()
            messages.success(request, 'Venta actualizada exitosamente.')
            return redirect('comercial:venta_detail', pk=pk)
    else:
        form = VentaForm(instance=venta)
    return render(request, 'comercial/ventas/form.html', {'form': form, 'title': 'Editar Venta'})


@login_required
def venta_delete(request, pk):
    venta = get_object_or_404(Venta, pk=pk)
    if request.method == 'POST':
        venta.delete()  # Eliminado l�gico
        messages.success(request, 'Venta eliminada exitosamente.')
        return redirect('comercial:ventas_list')
    return redirect('comercial:ventas_list')


@login_required
def venta_detail(request, pk):
    from django.utils import timezone
    venta = get_object_or_404(Venta.objects.select_related('cliente').prefetch_related('pagos'), pk=pk)
    pagos = venta.pagos.all().order_by('fecha_pago')
    total_pagado = venta.sena + sum(p.monto for p in pagos)
    total = venta.get_total_con_percepciones()
    porcentaje_cobrado = int((total_pagado / total * 100)) if total > 0 else 0
    dias_abierta = (timezone.now().date() - venta.created_at.date()).days

    context = {
        'venta': venta,
        'pagos': pagos,
        'total_pagado': total_pagado,
        'porcentaje_cobrado': porcentaje_cobrado,
        'dias_abierta': dias_abierta,
        'today': timezone.now().date(),
        'forzar_colocado': request.GET.get('forzar_colocado') == '1',
    }
    return render(request, 'comercial/ventas/detail.html', context)


@login_required
def duplicar_venta(request, pk):
    original = get_object_or_404(Venta, pk=pk, deleted_at__isnull=True)
    nueva = Venta.objects.create(
        numero_pedido=original.numero_pedido,
        cliente=original.cliente,
        valor_total=original.valor_total,
        sena=original.sena,
        con_factura=original.con_factura,
        forma_pago=original.forma_pago,
        observaciones=original.observaciones,
    )
    messages.success(request, f'Venta duplicada exitosamente. Nueva Venta #{nueva.pk}')
    return redirect('comercial:venta_edit', pk=nueva.pk)


@login_required
def guardar_nota_venta(request, pk):
    if request.method != 'POST':
        return JsonResponse({'error': 'Metodo no permitido'}, status=405)
    venta = get_object_or_404(Venta, pk=pk)
    try:
        venta.notas_internas = request.POST.get('nota', '')
        venta.save()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def registrar_pago(request, pk):
    from .models import Retencion
    venta = get_object_or_404(Venta, pk=pk)
    
    if request.method == 'POST':
        monto = request.POST.get('monto')
        fecha_pago = request.POST.get('fecha_pago')
        forma_pago = request.POST.get('forma_pago')
        con_factura = request.POST.get('con_factura') == 'true'
        numero_factura = request.POST.get('numero_factura', '')

        observaciones = request.POST.get('observaciones', '')
        fecha_factura = request.POST.get('fecha_factura', None)
        pago_en_dolares = request.POST.get('pago_en_dolares') == 'true'
        monto_usd = request.POST.get('monto_usd', '') or None
        cotizacion_usd = request.POST.get('cotizacion_usd', '') or None

        try:
            monto_decimal, monto_usd_decimal, cotizacion_usd_decimal = _resolver_monto_pago_venta(
                monto=monto,
                pago_en_dolares=pago_en_dolares,
                monto_usd=monto_usd,
                cotizacion_usd=cotizacion_usd,
            )

            if monto_decimal <= 0:
                messages.error(request, 'El monto debe ser mayor a 0')
                return redirect('comercial:venta_detail', pk=pk)

            if monto_decimal > venta.saldo:
                messages.error(request, f'El monto no puede ser mayor al saldo pendiente (${venta.saldo})')
                return redirect('comercial:venta_detail', pk=pk)

            pago = PagoVenta.objects.create(
                venta=venta,
                monto=monto_decimal,
                fecha_pago=fecha_pago,
                forma_pago=forma_pago,
                con_factura=con_factura,
                numero_factura=numero_factura,
                fecha_factura=fecha_factura if fecha_factura else None,
                observaciones=observaciones,
                pago_en_dolares=pago_en_dolares,
                monto_usd=monto_usd_decimal if pago_en_dolares else None,
                cotizacion_usd=cotizacion_usd_decimal if pago_en_dolares else None,
                created_by=request.user
            )
            
            # Procesar retenciones
            retencion_counts = request.POST.getlist('retencion_count')
            for count in retencion_counts:
                tipo = request.POST.get(f'retencion_tipo_{count}')
                importe = request.POST.get(f'retencion_importe_{count}')
                concepto = request.POST.get(f'retencion_concepto_{count}', '')
                comprobante = request.POST.get(f'retencion_comprobante_{count}', '')
                
                if tipo and importe:
                    importe_decimal = Decimal(importe)
                    if importe_decimal > 0:
                        Retencion.objects.create(
                            pago=pago,
                            tipo=tipo,
                            concepto=concepto,
                            numero_comprobante=comprobante,
                            importe_retenido=importe_decimal,
                            fecha_comprobante=fecha_pago
                        )
            
            # Recalcular saldo
            venta.save()

            messages.success(request, f'Pago de ${monto_decimal} registrado exitosamente')
            # Si el saldo quedo en 0 y el estado no es colocado, forzar cambio a colocado
            if venta.saldo <= 0 and venta.estado != 'colocado':
                return redirect(f"{reverse('comercial:venta_detail', kwargs={'pk': pk})}?forzar_colocado=1")
            return redirect('comercial:venta_detail', pk=pk)
            
        except Exception as e:
            messages.error(request, f'Error al registrar el pago: {str(e)}')
            return redirect('comercial:venta_detail', pk=pk)
    
    return redirect('comercial:venta_detail', pk=pk)


@login_required
def cambiar_estado_venta(request, pk):
    if request.method != 'POST':
        return JsonResponse({'error': 'Metodo no permitido'}, status=405)

    venta = get_object_or_404(Venta, pk=pk)
    nuevo_estado = request.POST.get('estado')

    estados_validos = [e[0] for e in Venta.ESTADO_CHOICES]
    if nuevo_estado not in estados_validos:
        return JsonResponse({'error': 'Estado no valido'}, status=400)

    venta.estado = nuevo_estado
    venta.save()

    return JsonResponse({'success': True, 'nuevo_estado': venta.get_estado_display()})


@login_required
def generar_pdf_venta(request, pk):
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    from reportlab.pdfgen import canvas
    
    venta = get_object_or_404(Venta.objects.select_related('cliente').prefetch_related('pagos'), pk=pk)
    pagos = venta.pagos.all().order_by('-fecha_pago')
    total_pagado = venta.sena + sum(p.monto for p in pagos)
    
    # Crear respuesta HTTP
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="venta_{venta.numero_pedido}.pdf"'
    
    # Crear PDF
    doc = SimpleDocTemplate(response, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    # Estilos personalizados
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=6,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=11,
        textColor=colors.HexColor('#64748b'),
        spaceAfter=20,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#1e293b'),
        spaceAfter=12,
        spaceBefore=20,
        fontName='Helvetica-Bold'
    )
    
    # Encabezado
    elements.append(Paragraph("AKUNA ABERTURAS", title_style))
    elements.append(Paragraph("Detalle de Venta", subtitle_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Informaci�n de la venta
    info_data = [
        ['Pedido N�:', venta.numero_pedido, 'Fecha:', venta.created_at.strftime('%d/%m/%Y')],
        ['Cliente:', f"{venta.cliente}", 'Estado:', venta.get_estado_display()],
    ]
    
    if venta.numero_factura:
        info_data.append(['Factura:', venta.get_numero_factura_display(), 'Tipo:', 'Blanco' if venta.con_factura else 'Negro'])
    
    info_table = Table(info_data, colWidths=[1.2*inch, 2.5*inch, 1*inch, 1.8*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f1f5f9')),
        ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#f1f5f9')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1e293b')),
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Resumen financiero
    elements.append(Paragraph("Resumen Financiero", heading_style))
    
    def format_currency(value):
        return f"${value:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    
    resumen_data = [
        ['Concepto', 'Monto'],
        ['Valor Total', format_currency(venta.valor_total)],
        ['Se�a Inicial', format_currency(venta.sena)],
        ['Pagos Adicionales', format_currency(sum(p.monto for p in pagos))],
        ['Total Pagado', format_currency(total_pagado)],
        ['Saldo Pendiente', format_currency(venta.saldo)],
    ]
    
    resumen_table = Table(resumen_data, colWidths=[4*inch, 2.5*inch])
    resumen_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#dbeafe')),
        ('BACKGROUND', (0, 4), (-1, 4), colors.HexColor('#dcfce7')),
        ('FONTNAME', (0, 4), (-1, 4), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 5), (-1, 5), colors.HexColor('#fef3c7') if venta.saldo > 0 else colors.HexColor('#dcfce7')),
        ('FONTNAME', (0, 5), (-1, 5), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 5), (-1, 5), 12),
    ]))
    elements.append(resumen_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Historial de pagos
    if pagos.exists() or venta.sena > 0:
        elements.append(Paragraph("Historial de Pagos", heading_style))
        
        pagos_data = [['Fecha', 'Concepto', 'Forma de Pago', 'N� Factura', 'Monto']]
        
        # Se�a inicial
        pagos_data.append([
            venta.created_at.strftime('%d/%m/%Y'),
            'Se�a Inicial',
            '-',
            venta.numero_factura or '-',
            format_currency(venta.sena)
        ])
        
        # Pagos adicionales
        for pago in pagos:
            pagos_data.append([
                pago.fecha_pago.strftime('%d/%m/%Y'),
                'Pago',
                pago.get_forma_pago_display(),
                pago.numero_factura or '-',
                format_currency(pago.monto)
            ])
        
        pagos_table = Table(pagos_data, colWidths=[1.1*inch, 1.5*inch, 1.3*inch, 1.3*inch, 1.3*inch])
        pagos_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (4, 0), (4, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ]))
        elements.append(pagos_table)
    
    # Pie de p�gina
    elements.append(Spacer(1, 0.5*inch))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#64748b'),
        alignment=TA_CENTER
    )
    elements.append(Paragraph(f"Documento generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}", footer_style))
    elements.append(Paragraph("Akuna Aberturas - Sistema de Gesti�n", footer_style))
    
    # Construir PDF
    doc.build(elements)
    return response


# CLIENTES
@login_required
def clientes_list(request):
    clientes = Cliente.objects.filter(deleted_at__isnull=True).all()
    
    # Filtros
    buscar = request.GET.get('q', '')
    if buscar:
        clientes = clientes.filter(
            Q(nombre__icontains=buscar) | 
            Q(apellido__icontains=buscar) | 
            Q(razon_social__icontains=buscar) |
            Q(localidad__icontains=buscar)
        )
    
    return render(request, 'comercial/clientes/list.html', {'clientes': clientes})


@login_required
def cliente_create(request):
    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():
            cliente = form.save()
            
            # Si es AJAX (desde modal), devolver JSON
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'cliente': {
                        'id': cliente.id,
                        'nombre': str(cliente)
                    }
                })
            
            messages.success(request, 'Cliente creado exitosamente.')
            return redirect('comercial:clientes_list')
        else:
            # Si es AJAX y hay errores
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'errors': form.errors
                }, status=400)
    else:
        form = ClienteForm()
    return render(request, 'comercial/clientes/form.html', {'form': form, 'title': 'Nuevo Cliente'})


@login_required
def cliente_edit(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    if request.method == 'POST':
        form = ClienteForm(request.POST, instance=cliente)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cliente actualizado exitosamente.')
            return redirect('comercial:clientes_list')
    else:
        form = ClienteForm(instance=cliente)
    return render(request, 'comercial/clientes/form.html', {'form': form, 'title': 'Editar Cliente'})


@login_required
def cliente_delete(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    if request.method == 'POST':
        # Contar ventas relacionadas
        ventas_relacionadas = Venta.objects.filter(cliente=cliente, deleted_at__isnull=True).count()
        
        # Eliminar el cliente (l�gico)
        cliente.delete()
        
        if ventas_relacionadas > 0:
            messages.success(request, f'Cliente eliminado. Hay {ventas_relacionadas} ventas que usaban este cliente.')
        else:
            messages.success(request, 'Cliente eliminado exitosamente.')
        
        return redirect('comercial:clientes_list')
    return redirect('comercial:clientes_list')


@login_required
def exportar_ventas_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    ventas = Venta.objects.filter(deleted_at__isnull=True).select_related('cliente').order_by('-created_at')

    # Mismos filtros que ventas_list
    estado = request.GET.get('estado', '')
    con_factura = request.GET.get('con_factura', '')
    buscar = request.GET.get('q', '')
    razon_social = request.GET.get('razon_social', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    con_saldo = request.GET.get('con_saldo', '')

    if estado:
        ventas = ventas.filter(estado=estado)
    if con_factura == 'si':
        ventas = ventas.filter(con_factura=True)
    elif con_factura == 'no':
        ventas = ventas.filter(con_factura=False)
    if buscar:
        ventas = ventas.filter(
            Q(numero_pedido__icontains=buscar) |
            Q(cliente__nombre__icontains=buscar) |
            Q(cliente__apellido__icontains=buscar) |
            Q(cliente__razon_social__icontains=buscar) |
            Q(numero_factura__icontains=buscar)
        )
    if razon_social:
        ventas = ventas.filter(cliente__razon_social__icontains=razon_social)
    if fecha_desde:
        ventas = ventas.filter(created_at__date__gte=fecha_desde)
    if fecha_hasta:
        ventas = ventas.filter(created_at__date__lte=fecha_hasta)
    if con_saldo == 'si':
        ventas = ventas.filter(saldo__gt=0)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"

    # Estilos
    header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    total_fill = PatternFill(start_color="dbeafe", end_color="dbeafe", fill_type="solid")
    total_font = Font(bold=True, size=11)
    thin = Side(style='thin', color='CBD5E1')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ['N� Pedido', 'Fecha', 'Cliente', 'Raz�n Social', 'Valor Total', 'Se�a', 'Saldo Pendiente', 'Estado', 'Tipo', 'N� Factura']
    col_widths = [14, 14, 28, 28, 16, 14, 18, 14, 10, 18]

    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 22

    for row_num, venta in enumerate(ventas, 2):
        datos = [
            venta.numero_pedido,
            venta.created_at.strftime('%d/%m/%Y'),
            str(venta.cliente),
            venta.cliente.razon_social or '',
            float(venta.valor_total),
            float(venta.sena),
            float(venta.saldo),
            venta.get_estado_display(),
            'Blanco' if venta.con_factura else 'Negro',
            venta.get_numero_factura_display(),
        ]
        for col_num, valor in enumerate(datos, 1):
            cell = ws.cell(row=row_num, column=col_num, value=valor)
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            if col_num in [5, 6, 7]:
                cell.number_format = '#,##0.00'

    # Fila de totales
    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=1, value='TOTALES').font = total_font
    ws.cell(row=total_row, column=5, value=sum(float(v.valor_total) for v in ventas)).number_format = '#,##0.00'
    ws.cell(row=total_row, column=6, value=sum(float(v.sena) for v in ventas)).number_format = '#,##0.00'
    ws.cell(row=total_row, column=7, value=sum(float(v.saldo) for v in ventas)).number_format = '#,##0.00'
    for col in range(1, 11):
        ws.cell(row=total_row, column=col).fill = total_fill
        ws.cell(row=total_row, column=col).font = total_font
        ws.cell(row=total_row, column=col).border = border

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="ventas.xlsx"'
    wb.save(response)
    return response


@login_required
def cliente_detail(request, pk):
    from facturacion.models import Factura
    from django.db.models import Sum, Count
    from django.db.models.functions import TruncMonth
    from datetime import datetime, timedelta
    import json
    from django.core.serializers.json import DjangoJSONEncoder

    cliente = get_object_or_404(Cliente, pk=pk, deleted_at__isnull=True)

    ventas = Venta.objects.filter(
        cliente=cliente, deleted_at__isnull=True
    ).prefetch_related('pagos', 'percepciones').order_by('-created_at')

    # KPIs
    kpis = ventas.aggregate(
        total_comprado=Sum('valor_total'),
        saldo_pendiente=Sum('saldo'),
        cantidad_ventas=Count('id'),
    )
    total_comprado = kpis['total_comprado'] or 0
    saldo_pendiente = kpis['saldo_pendiente'] or 0
    cantidad_ventas = kpis['cantidad_ventas'] or 0

    # Pagos recibidos
    pagos = PagoVenta.objects.filter(
        venta__cliente=cliente,
        venta__deleted_at__isnull=True,
    ).select_related('venta').order_by('-fecha_pago')
    total_cobrado = pagos.aggregate(total=Sum('monto'))['total'] or 0

    # Facturas electr�nicas
    facturas = Factura.objects.filter(cliente=cliente).select_related('venta', 'punto_venta').order_by('-fecha')
    cantidad_facturas = facturas.count()

    # Gr�fico 1: ventas por mes (�ltimos 12 meses)
    hace_12_meses = datetime.now() - timedelta(days=365)
    ventas_por_mes = (
        ventas.filter(created_at__gte=hace_12_meses)
        .order_by()
        .annotate(mes=TruncMonth('created_at'))
        .values('mes')
        .annotate(total=Sum('valor_total'))
        .order_by('mes')
    )
    meses_labels = [v['mes'].strftime('%b %Y') for v in ventas_por_mes]
    meses_data = [float(v['total']) for v in ventas_por_mes]

    # Gr�fico 2: distribuci�n por estado
    ESTADO_DISPLAY = {'pendiente': 'Pendiente', 'entregado': 'Entregado', 'colocado': 'Colocado'}
    estados_qs = ventas.order_by().values('estado').annotate(cantidad=Count('id'))
    estados_labels = [ESTADO_DISPLAY.get(e['estado'], e['estado']) for e in estados_qs]
    estados_data = [e['cantidad'] for e in estados_qs]

    context = {
        'cliente': cliente,
        'ventas': ventas,
        'pagos': pagos[:30],
        'facturas': facturas,
        'total_comprado': total_comprado,
        'saldo_pendiente': saldo_pendiente,
        'cantidad_ventas': cantidad_ventas,
        'cantidad_facturas': cantidad_facturas,
        'total_cobrado': total_cobrado,
        'meses_labels': json.dumps(meses_labels, cls=DjangoJSONEncoder),
        'meses_data': json.dumps(meses_data, cls=DjangoJSONEncoder),
        'estados_labels': json.dumps(estados_labels, cls=DjangoJSONEncoder),
        'estados_data': json.dumps(estados_data, cls=DjangoJSONEncoder),
    }
    return render(request, 'comercial/clientes/detail.html', context)


# COMPRAS
@login_required
def compras_list(request):
    from django.core.paginator import Paginator
    
    compras = Compra.objects.filter(deleted_at__isnull=True).select_related('cuenta', 'cuenta__tipo_cuenta').all()
    
    # Filtros
    tipo_cuenta = request.GET.get('tipo_cuenta', '')
    con_factura = request.GET.get('con_factura', '')
    buscar = request.GET.get('q', '')
    
    if tipo_cuenta:
        compras = compras.filter(cuenta__tipo_cuenta_id=tipo_cuenta)
    
    if con_factura == 'si':
        compras = compras.filter(con_factura=True)
    elif con_factura == 'no':
        compras = compras.filter(con_factura=False)
    
    if buscar:
        compras = compras.filter(
            Q(numero_pedido__icontains=buscar) |
            Q(cuenta__nombre__icontains=buscar) |
            Q(numero_factura__icontains=buscar)
        )
    
    # Ordenar por fecha de pago descendente
    compras = compras.order_by('-fecha_pago')
    
    # Paginaci�n
    paginator = Paginator(compras, 20)  # 20 items por p�gina
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'compras': page_obj,
        'page_obj': page_obj,
        'tipos_cuenta': TipoCuenta.objects.filter(activo=True),
        'filtro_tipo': tipo_cuenta,
        'filtro_factura': con_factura,
        'buscar': buscar,
        'titulo': 'Gastos'
    }
    return render(request, 'comercial/compras/list.html', context)


@login_required
def compra_create(request):
    if request.method == 'POST':
        form = CompraForm(request.POST)
        if form.is_valid():
            compra = form.save(commit=False)
            compra.created_by = request.user
            compra.save()
            messages.success(request, 'Gasto registrado exitosamente.')
            return redirect('comercial:compra_detail', pk=compra.pk)
    else:
        form = CompraForm()
    
    # Pasar tipos de cuenta al contexto
    context = {
        'form': form,
        'title': 'Nuevo Gasto',
        'tipos_cuenta': TipoCuenta.objects.filter(activo=True, deleted_at__isnull=True)
    }
    return render(request, 'comercial/compras/form.html', context)


@login_required
def compra_edit(request, pk):
    compra = get_object_or_404(Compra, pk=pk)
    if request.method == 'POST':
        form = CompraForm(request.POST, instance=compra)
        if form.is_valid():
            form.save()
            messages.success(request, 'Gasto actualizado exitosamente.')
            return redirect('comercial:compra_detail', pk=pk)
        # Si hay errores, pre-seleccionar el tipo de cuenta del POST
        tipo_cuenta_id = request.POST.get('tipo_cuenta_filter')
        if tipo_cuenta_id:
            form.fields['tipo_cuenta_filter'].initial = tipo_cuenta_id
    else:
        form = CompraForm(instance=compra)
        # Pre-seleccionar el tipo de cuenta
        if compra.cuenta:
            form.fields['tipo_cuenta_filter'].initial = compra.cuenta.tipo_cuenta_id
    
    context = {
        'form': form,
        'title': 'Editar Gasto',
        'tipos_cuenta': TipoCuenta.objects.filter(activo=True, deleted_at__isnull=True),
        'compra': compra
    }
    return render(request, 'comercial/compras/form.html', context)


@login_required
def compra_delete(request, pk):
    compra = get_object_or_404(Compra, pk=pk)
    if request.method == 'POST':
        compra.delete()
        messages.success(request, 'Gasto eliminado exitosamente.')
        return redirect('comercial:compras_list')
    return redirect('comercial:compras_list')


@login_required
def compra_detail(request, pk):
    from django.utils import timezone
    compra = get_object_or_404(Compra.objects.select_related('cuenta', 'cuenta__tipo_cuenta', 'tipo_gasto').prefetch_related('pagos_compra'), pk=pk, deleted_at__isnull=True)
    pagos = compra.pagos_compra.all().order_by('fecha_pago')
    total_pagado = compra.sena + sum(p.monto for p in pagos)
    porcentaje_pagado = int((total_pagado / compra.valor_total * 100)) if compra.valor_total > 0 else 0
    dias_abierta = (timezone.now().date() - compra.created_at.date()).days

    context = {
        'compra': compra,
        'pagos': pagos,
        'total_pagado': total_pagado,
        'porcentaje_pagado': porcentaje_pagado,
        'dias_abierta': dias_abierta,
        'today': timezone.now().date(),
        'permite_pago_adicional': compra.estado == 'pendiente',
    }
    return render(request, 'comercial/compras/detail.html', context)


@login_required
def registrar_pago_compra(request, pk):
    compra = get_object_or_404(Compra, pk=pk)
    if request.method == 'POST':
        monto = request.POST.get('monto')
        fecha_pago = request.POST.get('fecha_pago')
        forma_pago = request.POST.get('forma_pago')
        con_factura = request.POST.get('con_factura') == 'true'
        numero_factura = request.POST.get('numero_factura', '')
        observaciones = request.POST.get('observaciones', '')
        es_proveedor = compra.es_proveedor

        try:
            monto_decimal = Decimal(monto)
            if monto_decimal <= 0:
                messages.error(request, 'El monto debe ser mayor a 0')
                return redirect('comercial:compra_detail', pk=pk)
            if not es_proveedor and monto_decimal > compra.saldo:
                messages.error(request, f'El monto no puede exceder el saldo pendiente (${compra.saldo})')
                return redirect('comercial:compra_detail', pk=pk)

            PagoCompra.objects.create(
                compra=compra,
                monto=monto_decimal,
                fecha_pago=fecha_pago,
                forma_pago=forma_pago,
                con_factura=con_factura,
                numero_factura=numero_factura,
                observaciones=observaciones,
                created_by=request.user
            )
            compra.save()
            messages.success(request, f'Pago de ${monto_decimal} registrado exitosamente.')

            if compra.saldo <= 0 and compra.estado == 'pendiente':
                compra.estado = 'pagado'
                compra.save()
                messages.success(request, 'La compra fue marcada como pagada.')

            if es_proveedor and compra.saldo < 0:
                messages.info(request, f'El proveedor quedó con saldo a favor de ${abs(compra.saldo)}.')

            return redirect('comercial:compra_detail', pk=pk)
        except Exception as e:
            messages.error(request, f'Error al registrar pago: {str(e)}')
            return redirect('comercial:compra_detail', pk=pk)
    return redirect('comercial:compra_detail', pk=pk)


@login_required
def editar_pago_compra(request, pk):
    if request.method == 'POST':
        import json
        pago = get_object_or_404(PagoCompra, pk=pk)
        try:
            data = json.loads(request.body)
            pago_en_dolares = data.get('pago_en_dolares', False)
            monto, monto_usd, cotizacion_usd = _resolver_monto_pago_venta(
                monto=data.get('monto'),
                pago_en_dolares=pago_en_dolares,
                monto_usd=data.get('monto_usd'),
                cotizacion_usd=data.get('cotizacion_usd'),
            )
            fecha_pago_str = data.get('fecha_pago')
            forma_pago = data.get('forma_pago')
            con_factura = data.get('con_factura', True)
            numero_factura = data.get('numero_factura', '')
            observaciones = data.get('observaciones', '')

            if monto <= 0:
                return JsonResponse({'success': False, 'error': 'El monto debe ser mayor a 0'}, status=400)

            pago.monto = monto
            pago.fecha_pago = datetime.strptime(fecha_pago_str, '%Y-%m-%d').date()
            pago.forma_pago = forma_pago
            pago.con_factura = con_factura
            pago.numero_factura = numero_factura
            pago.observaciones = observaciones
            pago.save()

            pago.compra.save()

            return JsonResponse({
                'success': True,
                'saldo': float(pago.compra.saldo)
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)


@login_required
def eliminar_pago_compra(request, pk):
    if request.method == 'POST':
        pago = get_object_or_404(PagoCompra, pk=pk)
        compra = pago.compra
        try:
            pago.delete()
            compra.save()
            if compra.saldo > 0 and compra.estado == 'pagado':
                compra.estado = 'pendiente'
                compra.save()
            return JsonResponse({
                'success': True,
                'saldo': float(compra.saldo)
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)


@login_required
def guardar_nota_compra(request, pk):
    if request.method == 'POST':
        compra = get_object_or_404(Compra, pk=pk)
        nota = request.POST.get('nota', '')
        Compra.objects.filter(pk=pk).update(notas_internas=nota)
        return JsonResponse({'success': True})
    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)


# CUENTAS
@login_required
def cuentas_list(request):
    cuentas = Cuenta.objects.filter(deleted_at__isnull=True).select_related('tipo_cuenta').all()
    
    # Filtros
    buscar = request.GET.get('q', '')
    tipo_cuenta_id = request.GET.get('tipo_cuenta', '')
    estado = request.GET.get('estado', '')
    
    if buscar:
        cuentas = cuentas.filter(
            Q(nombre__icontains=buscar) | Q(razon_social__icontains=buscar)
        )
    
    if tipo_cuenta_id:
        cuentas = cuentas.filter(tipo_cuenta_id=tipo_cuenta_id)
    
    if estado == 'activo':
        cuentas = cuentas.filter(activo=True)
    elif estado == 'inactivo':
        cuentas = cuentas.filter(activo=False)
    
    tipos_cuenta_filtro = TipoCuenta.objects.filter(activo=True, deleted_at__isnull=True)
    
    return render(request, 'comercial/cuentas/list.html', {
        'cuentas': cuentas,
        'tipos_cuenta_filtro': tipos_cuenta_filtro
    })


@login_required
def cuenta_create(request):
    if request.method == 'POST':
        form = CuentaForm(request.POST)
        if form.is_valid():
            cuenta = form.save()
            
            # Si es AJAX (desde modal), devolver JSON
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'cuenta': {
                        'id': cuenta.id,
                        'nombre': str(cuenta)
                    }
                })
            
            messages.success(request, 'Cuenta creada exitosamente.')
            return redirect('comercial:cuentas_list')
        else:
            # Si es AJAX y hay errores
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'errors': form.errors
                }, status=400)
    else:
        form = CuentaForm()
    return render(request, 'comercial/cuentas/form.html', {'form': form, 'title': 'Nueva Cuenta'})


@login_required
def cuenta_edit(request, pk):
    cuenta = get_object_or_404(Cuenta, pk=pk)
    if request.method == 'POST':
        form = CuentaForm(request.POST, instance=cuenta)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cuenta actualizada exitosamente.')
            return redirect('comercial:cuentas_list')
    else:
        form = CuentaForm(instance=cuenta)
    return render(request, 'comercial/cuentas/form.html', {'form': form, 'title': 'Editar Cuenta'})


@login_required
def cuenta_delete(request, pk):
    cuenta = get_object_or_404(Cuenta, pk=pk)
    if request.method == 'POST':
        # Contar compras relacionadas
        compras_relacionadas = Compra.objects.filter(cuenta=cuenta, deleted_at__isnull=True).count()
        
        # Eliminar la cuenta (l�gico)
        cuenta.delete()
        
        if compras_relacionadas > 0:
            messages.success(request, f'Cuenta eliminada. Hay {compras_relacionadas} compras que usaban esta cuenta.')
        else:
            messages.success(request, 'Cuenta eliminada exitosamente.')
        
        return redirect('comercial:cuentas_list')
    return redirect('comercial:cuentas_list')


# TIPOS DE CUENTA
@login_required
def tipos_cuenta_list(request):
    tipos = TipoCuenta.objects.filter(deleted_at__isnull=True).all()
    
    # Filtros
    buscar = request.GET.get('q', '')
    estado = request.GET.get('estado', '')
    
    if buscar:
        tipos = tipos.filter(
            Q(descripcion__icontains=buscar) | Q(tipo__icontains=buscar)
        )
    
    if estado == 'activo':
        tipos = tipos.filter(activo=True)
    elif estado == 'inactivo':
        tipos = tipos.filter(activo=False)
    
    return render(request, 'comercial/tipos_cuenta/list.html', {'tipos': tipos})


@login_required
def tipo_cuenta_create(request):
    if request.method == 'POST':
        tipo = request.POST.get('tipo')
        descripcion = request.POST.get('descripcion')
        activo = request.POST.get('activo') == 'on'
        
        if tipo and descripcion:
            # Verificar si existe un registro eliminado
            tipo_existente = TipoCuenta.objects.filter(tipo=tipo).first()
            if tipo_existente:
                # Reactivar el registro eliminado
                tipo_existente.descripcion = descripcion
                tipo_existente.activo = activo
                tipo_existente.deleted_at = None
                tipo_existente.save()
                messages.success(request, 'Tipo de cuenta reactivado exitosamente.')
            else:
                # Crear nuevo registro
                TipoCuenta.objects.create(
                    tipo=tipo,
                    descripcion=descripcion,
                    activo=activo
                )
                messages.success(request, 'Tipo de cuenta creado exitosamente.')
            return redirect('comercial:tipos_cuenta_list')
    
    return render(request, 'comercial/tipos_cuenta/form.html', {'title': 'Nuevo Tipo de Cuenta'})


@login_required
def tipo_cuenta_edit(request, pk):
    tipo = get_object_or_404(TipoCuenta, pk=pk)
    
    if request.method == 'POST':
        tipo.tipo = request.POST.get('tipo')
        tipo.descripcion = request.POST.get('descripcion')
        tipo.activo = request.POST.get('activo') == 'on'
        tipo.save()
        messages.success(request, 'Tipo de cuenta actualizado exitosamente.')
        return redirect('comercial:tipos_cuenta_list')
    
    return render(request, 'comercial/tipos_cuenta/form.html', {'tipo': tipo, 'title': 'Editar Tipo de Cuenta'})


@login_required
def tipo_cuenta_delete(request, pk):
    tipo = get_object_or_404(TipoCuenta, pk=pk)
    if request.method == 'POST':
        # Desactivar cuentas relacionadas
        cuentas_afectadas = Cuenta.objects.filter(tipo_cuenta=tipo, deleted_at__isnull=True)
        for cuenta in cuentas_afectadas:
            cuenta.delete()  # Eliminado l�gico
        
        # Desactivar tipos de gasto relacionados
        tipos_gasto_afectados = TipoGasto.objects.filter(tipo_cuenta=tipo, deleted_at__isnull=True)
        for tipo_gasto in tipos_gasto_afectados:
            tipo_gasto.delete()  # Eliminado l�gico
        
        # Eliminar el tipo de cuenta
        tipo.delete()
        messages.success(request, f'Tipo de cuenta eliminado. Se desactivaron {cuentas_afectadas.count()} cuentas y {tipos_gasto_afectados.count()} tipos de gasto.')
        return redirect('comercial:tipos_cuenta_list')
    return redirect('comercial:tipos_cuenta_list')


# TIPOS DE GASTO
@login_required
def tipos_gasto_list(request):
    tipos = TipoGasto.objects.filter(deleted_at__isnull=True).select_related('tipo_cuenta').all()
    
    # Filtros
    buscar = request.GET.get('q', '')
    tipo_cuenta_id = request.GET.get('tipo_cuenta', '')
    estado = request.GET.get('estado', '')
    
    if buscar:
        tipos = tipos.filter(
            Q(nombre__icontains=buscar) | Q(descripcion__icontains=buscar)
        )
    
    if tipo_cuenta_id:
        tipos = tipos.filter(tipo_cuenta_id=tipo_cuenta_id)
    
    if estado == 'activo':
        tipos = tipos.filter(activo=True)
    elif estado == 'inactivo':
        tipos = tipos.filter(activo=False)
    
    tipos_cuenta_filtro = TipoCuenta.objects.filter(activo=True, deleted_at__isnull=True)
    
    return render(request, 'comercial/tipos_gasto/list.html', {
        'tipos': tipos,
        'tipos_cuenta_filtro': tipos_cuenta_filtro
    })


@login_required
def tipo_gasto_create(request):
    if request.method == 'POST':
        tipo_cuenta_id = request.POST.get('tipo_cuenta')
        nombre = request.POST.get('nombre')
        descripcion = request.POST.get('descripcion', '')
        
        if tipo_cuenta_id and nombre:
            TipoGasto.objects.create(
                tipo_cuenta_id=tipo_cuenta_id,
                nombre=nombre,
                descripcion=descripcion
            )
            messages.success(request, 'Tipo de gasto creado exitosamente.')
            return redirect('comercial:tipos_gasto_list')
    
    tipos_cuenta = TipoCuenta.objects.filter(activo=True, deleted_at__isnull=True)
    return render(request, 'comercial/tipos_gasto/form.html', {'tipos_cuenta': tipos_cuenta, 'title': 'Nuevo Tipo de Gasto'})


@login_required
def tipo_gasto_edit(request, pk):
    tipo = get_object_or_404(TipoGasto, pk=pk)
    
    if request.method == 'POST':
        tipo_cuenta_id = request.POST.get('tipo_cuenta')
        nombre = request.POST.get('nombre')
        descripcion = request.POST.get('descripcion', '')
        
        if tipo_cuenta_id and nombre:
            tipo.tipo_cuenta_id = tipo_cuenta_id
            tipo.nombre = nombre
            tipo.descripcion = descripcion
            tipo.save()
            messages.success(request, 'Tipo de gasto actualizado exitosamente.')
            return redirect('comercial:tipos_gasto_list')
    
    tipos_cuenta = TipoCuenta.objects.filter(activo=True, deleted_at__isnull=True)
    return render(request, 'comercial/tipos_gasto/form.html', {'tipos_cuenta': tipos_cuenta, 'tipo': tipo, 'title': 'Editar Tipo de Gasto'})


@login_required
def tipo_gasto_delete(request, pk):
    tipo = get_object_or_404(TipoGasto, pk=pk)
    if request.method == 'POST':
        # Contar compras relacionadas
        compras_relacionadas = Compra.objects.filter(tipo_gasto=tipo, deleted_at__isnull=True).count()
        
        # Eliminar el tipo de gasto (l�gico)
        tipo.delete()
        
        if compras_relacionadas > 0:
            messages.success(request, f'Tipo de gasto eliminado. Hay {compras_relacionadas} compras que usaban este tipo.')
        else:
            messages.success(request, 'Tipo de gasto eliminado exitosamente.')
        
        return redirect('comercial:tipos_gasto_list')
    return redirect('comercial:tipos_gasto_list')


# REPORTES
def construir_reporte_ventas(
    fecha_desde=None,
    fecha_hasta=None,
    cliente_filtro=None,
    razon_social_filtro=None,
    estado_venta_filtro=None,
    tipo_factura_filtro=None,
):
    ventas_query = Venta.objects.filter(deleted_at__isnull=True).select_related('cliente').prefetch_related('percepciones')

    if fecha_desde:
        ventas_query = ventas_query.filter(fecha_pago__gte=fecha_desde)
    if fecha_hasta:
        ventas_query = ventas_query.filter(fecha_pago__lte=fecha_hasta)
    if cliente_filtro:
        ventas_query = ventas_query.filter(cliente__in=cliente_filtro)
    if razon_social_filtro:
        ventas_query = ventas_query.filter(cliente__razon_social__in=razon_social_filtro)
    if estado_venta_filtro:
        ventas_query = ventas_query.filter(estado__in=estado_venta_filtro)
    if tipo_factura_filtro:
        if 'blanco' in tipo_factura_filtro and 'negro' not in tipo_factura_filtro:
            ventas_query = ventas_query.filter(con_factura=True)
        elif 'negro' in tipo_factura_filtro and 'blanco' not in tipo_factura_filtro:
            ventas_query = ventas_query.filter(con_factura=False)

    ventas = []
    for venta in ventas_query.order_by('-fecha_pago', '-created_at'):
        ventas.append({
            'fecha': venta.fecha_pago or venta.created_at.date(),
            'pedido': venta.numero_pedido,
            'numero_factura': venta.numero_factura or '-',
            'factura_id': venta.id if venta.numero_factura else None,
            'cliente': str(venta.cliente),
            'razon_social': venta.cliente.razon_social or '-',
            'forma_pago': venta.get_forma_pago_display() if venta.forma_pago else '-',
            'monto': venta.valor_total,
            'tipo': 'Blanco' if venta.con_factura else 'Negro',
            'venta_id': venta.id,
        })

    return ventas


def construir_reporte_cobranzas(
    fecha_desde=None,
    fecha_hasta=None,
    cliente_filtro=None,
    razon_social_filtro=None,
    estado_venta_filtro=None,
    tipo_factura_filtro=None,
    numero_factura_filtro=None,
):
    cobranzas = []

    ventas_query = Venta.objects.filter(deleted_at__isnull=True, sena__gt=0).select_related('cliente')

    if fecha_desde:
        ventas_query = ventas_query.filter(created_at__date__gte=fecha_desde)
    if fecha_hasta:
        ventas_query = ventas_query.filter(created_at__date__lte=fecha_hasta)
    if cliente_filtro:
        ventas_query = ventas_query.filter(cliente__in=cliente_filtro)
    if razon_social_filtro:
        ventas_query = ventas_query.filter(cliente__razon_social__in=razon_social_filtro)
    if estado_venta_filtro:
        ventas_query = ventas_query.filter(estado__in=estado_venta_filtro)
    if tipo_factura_filtro:
        if 'blanco' in tipo_factura_filtro and 'negro' not in tipo_factura_filtro:
            ventas_query = ventas_query.filter(con_factura=True)
        elif 'negro' in tipo_factura_filtro and 'blanco' not in tipo_factura_filtro:
            ventas_query = ventas_query.filter(con_factura=False)
    if numero_factura_filtro:
        ventas_query = ventas_query.filter(numero_factura__icontains=numero_factura_filtro)

    for venta in ventas_query:
        cobranzas.append({
            'fecha': venta.created_at.date(),
            'pedido': venta.numero_pedido,
            'numero_factura': venta.numero_factura or '-',
            'cliente': str(venta.cliente),
            'razon_social': venta.cliente.razon_social or '-',
            'concepto': 'Seña inicial',
            'forma_pago': venta.get_forma_pago_display() if venta.forma_pago else '-',
            'monto': venta.sena,
            'tipo': 'Blanco' if venta.con_factura else 'Negro',
            'venta_id': venta.id,
            'pago_en_dolares': False,
            'monto_usd': None,
            'cotizacion_usd': None,
        })

    pagos_query = PagoVenta.objects.filter(venta__deleted_at__isnull=True).select_related('venta', 'venta__cliente')

    if fecha_desde:
        pagos_query = pagos_query.filter(fecha_pago__gte=fecha_desde)
    if fecha_hasta:
        pagos_query = pagos_query.filter(fecha_pago__lte=fecha_hasta)
    if cliente_filtro:
        pagos_query = pagos_query.filter(venta__cliente__in=cliente_filtro)
    if razon_social_filtro:
        pagos_query = pagos_query.filter(venta__cliente__razon_social__in=razon_social_filtro)
    if estado_venta_filtro:
        pagos_query = pagos_query.filter(venta__estado__in=estado_venta_filtro)
    if tipo_factura_filtro:
        if 'blanco' in tipo_factura_filtro and 'negro' not in tipo_factura_filtro:
            pagos_query = pagos_query.filter(con_factura=True)
        elif 'negro' in tipo_factura_filtro and 'blanco' not in tipo_factura_filtro:
            pagos_query = pagos_query.filter(con_factura=False)
    if numero_factura_filtro:
        pagos_query = pagos_query.filter(
            Q(numero_factura__icontains=numero_factura_filtro) |
            Q(venta__numero_factura__icontains=numero_factura_filtro)
        )

    for pago in pagos_query.order_by('-fecha_pago', '-created_at'):
        cobranzas.append({
            'fecha': pago.fecha_pago,
            'pedido': pago.venta.numero_pedido,
            'numero_factura': pago.numero_factura or pago.venta.numero_factura or '-',
            'cliente': str(pago.venta.cliente),
            'razon_social': pago.venta.cliente.razon_social or '-',
            'concepto': 'Pago',
            'forma_pago': pago.get_forma_pago_display(),
            'monto': pago.monto,
            'tipo': 'Blanco' if pago.con_factura else 'Negro',
            'venta_id': pago.venta.id,
            'pago_en_dolares': pago.pago_en_dolares,
            'monto_usd': pago.monto_usd,
            'cotizacion_usd': pago.cotizacion_usd,
        })

    cobranzas.sort(key=lambda item: item['fecha'], reverse=True)
    return cobranzas


@login_required
def reportes(request):
    form = ReporteForm()
    
    # Obtener filtros
    fecha_desde = None
    fecha_hasta = None
    cliente_filtro = None
    razon_social_filtro = None
    estado_venta_filtro = None
    tipo_factura_filtro = None
    
    if request.method == 'POST':
        form = ReporteForm(request.POST)
        if form.is_valid():
            fecha_desde = form.cleaned_data.get('fecha_desde')
            fecha_hasta = form.cleaned_data.get('fecha_hasta')
            cliente_filtro = form.cleaned_data.get('cliente')
            razon_social_filtro = form.cleaned_data.get('razon_social')
            estado_venta_filtro = form.cleaned_data.get('estado_venta')
            tipo_factura_filtro = form.cleaned_data.get('tipo_factura')
            
            request.session['reporte_filtros'] = {
                'fecha_desde': fecha_desde.isoformat() if fecha_desde else None,
                'fecha_hasta': fecha_hasta.isoformat() if fecha_hasta else None,
                'cliente_id': [c.id for c in cliente_filtro] if cliente_filtro else None,
                'razon_social': list(razon_social_filtro) if razon_social_filtro else None,
                'estado_venta': list(estado_venta_filtro) if estado_venta_filtro else None,
                'tipo_factura': list(tipo_factura_filtro) if tipo_factura_filtro else None,
            }
    
    ventas_reporte = construir_reporte_ventas(
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        cliente_filtro=cliente_filtro,
        razon_social_filtro=razon_social_filtro,
        estado_venta_filtro=estado_venta_filtro,
        tipo_factura_filtro=tipo_factura_filtro,
    )
    
    # Calcular totales
    total_blanco = sum(i['monto'] for i in ventas_reporte if i['tipo'] == 'Blanco')
    total_negro = sum(i['monto'] for i in ventas_reporte if i['tipo'] == 'Negro')
    total = total_blanco + total_negro
    
    cantidad_blanco = len([i for i in ventas_reporte if i['tipo'] == 'Blanco'])
    cantidad_negro = len([i for i in ventas_reporte if i['tipo'] == 'Negro'])
    
    reporte_data = {
        'ventas': {
            'total_blanco': total_blanco,
            'total_negro': total_negro,
            'total': total,
            'cantidad_blanco': cantidad_blanco,
            'cantidad_negro': cantidad_negro,
            'cantidad': len(ventas_reporte),
            'lista': ventas_reporte
        }
    }
    
    context = {
        'form': form,
        'reporte_data': reporte_data,
    }
    return render(request, 'comercial/reportes/reportes.html', context)


@login_required
def reportes_cobranzas(request):
    form = ReporteForm()

    fecha_desde = None
    fecha_hasta = None
    cliente_filtro = None
    razon_social_filtro = None
    estado_venta_filtro = None
    tipo_factura_filtro = None
    numero_factura_filtro = None

    if request.method == 'POST':
        form = ReporteForm(request.POST)
        if form.is_valid():
            fecha_desde = form.cleaned_data.get('fecha_desde')
            fecha_hasta = form.cleaned_data.get('fecha_hasta')
            cliente_filtro = form.cleaned_data.get('cliente')
            razon_social_filtro = form.cleaned_data.get('razon_social')
            estado_venta_filtro = form.cleaned_data.get('estado_venta')
            tipo_factura_filtro = form.cleaned_data.get('tipo_factura')
            numero_factura_filtro = form.cleaned_data.get('numero_factura')

    cobranzas = construir_reporte_cobranzas(
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        cliente_filtro=cliente_filtro,
        razon_social_filtro=razon_social_filtro,
        estado_venta_filtro=estado_venta_filtro,
        tipo_factura_filtro=tipo_factura_filtro,
        numero_factura_filtro=numero_factura_filtro,
    )

    total_blanco = sum(item['monto'] for item in cobranzas if item['tipo'] == 'Blanco')
    total_negro = sum(item['monto'] for item in cobranzas if item['tipo'] == 'Negro')
    total = total_blanco + total_negro

    reporte_data = {
        'cobranzas': {
            'total_blanco': total_blanco,
            'total_negro': total_negro,
            'total': total,
            'cantidad_blanco': len([item for item in cobranzas if item['tipo'] == 'Blanco']),
            'cantidad_negro': len([item for item in cobranzas if item['tipo'] == 'Negro']),
            'cantidad': len(cobranzas),
            'lista': cobranzas,
        }
    }

    return render(request, 'comercial/reportes/reportes_cobranzas.html', {
        'form': form,
        'reporte_data': reporte_data,
    })


@login_required
def get_tipos_gasto_by_cuenta(request):
    cuenta_id = request.GET.get('cuenta_id')
    if cuenta_id:
        cuenta = Cuenta.objects.filter(id=cuenta_id).first()
        if cuenta:
            tipos_gasto = TipoGasto.objects.filter(
                tipo_cuenta=cuenta.tipo_cuenta,
                activo=True,
                deleted_at__isnull=True
            ).values('id', 'nombre')
            return JsonResponse(list(tipos_gasto), safe=False)
    return JsonResponse([], safe=False)


@login_required
def get_cuentas_by_tipo(request):
    tipo_id = request.GET.get('tipo_id')
    cuentas = Cuenta.objects.filter(tipo_cuenta_id=tipo_id, activo=True).values('id', 'nombre')
    return JsonResponse(list(cuentas), safe=False)


@login_required
def get_clientes_list(request):
    clientes = Cliente.objects.filter(deleted_at__isnull=True).values('id', 'nombre', 'apellido')
    return JsonResponse(list(clientes), safe=False)


@login_required
def editar_pago(request, pk):
    import json
    from datetime import datetime
    
    if request.method == 'POST':
        pago = get_object_or_404(PagoVenta, pk=pk)
        
        try:
            data = json.loads(request.body)

            pago_en_dolares = data.get('pago_en_dolares', False)
            monto, monto_usd, cotizacion_usd = _resolver_monto_pago_venta(
                monto=data.get('monto'),
                pago_en_dolares=pago_en_dolares,
                monto_usd=data.get('monto_usd'),
                cotizacion_usd=data.get('cotizacion_usd'),
            )
            fecha_pago_str = data.get('fecha_pago')
            forma_pago = data.get('forma_pago')
            con_factura = data.get('con_factura', True)
            numero_factura = data.get('numero_factura', '')
            observaciones = data.get('observaciones', '')

            if monto <= 0:
                return JsonResponse({'success': False, 'error': 'El monto debe ser mayor a 0'}, status=400)

            # Convertir fecha string a objeto date
            if isinstance(fecha_pago_str, str):
                fecha_pago = datetime.strptime(fecha_pago_str, '%Y-%m-%d').date()
            else:
                fecha_pago = fecha_pago_str

            pago.monto = monto
            pago.fecha_pago = fecha_pago
            pago.forma_pago = forma_pago
            pago.con_factura = con_factura
            pago.numero_factura = numero_factura
            pago.observaciones = observaciones
            pago.pago_en_dolares = pago_en_dolares
            pago.monto_usd = monto_usd if pago_en_dolares else None
            pago.cotizacion_usd = cotizacion_usd if pago_en_dolares else None
            pago.save()

            # Recalcular saldo de la venta
            pago.venta.save()

            return JsonResponse({
                'success': True,
                'pago': {
                    'id': pago.id,
                    'monto': float(pago.monto),
                    'fecha_pago': pago.fecha_pago.strftime('%d/%m/%Y'),
                    'forma_pago': pago.get_forma_pago_display(),
                    'con_factura': pago.con_factura,
                    'numero_factura': pago.numero_factura or '-',
                    'observaciones': pago.observaciones or '-',
                    'pago_en_dolares': pago.pago_en_dolares,
                    'monto_usd': float(pago.monto_usd) if pago.monto_usd else None,
                    'cotizacion_usd': float(pago.cotizacion_usd) if pago.cotizacion_usd else None,
                },
                'saldo': float(pago.venta.saldo)
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    
    return JsonResponse({'success': False, 'error': 'Metodo no permitido'}, status=405)


@login_required
def exportar_reporte_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse
    
    # Recuperar filtros de la sesi�n
    filtros = request.session.get('reporte_filtros', {})
    
    cliente_ids = filtros.get('cliente_id')
    cliente_filtro = Cliente.objects.filter(id__in=cliente_ids) if cliente_ids else None
    ventas_reporte = construir_reporte_ventas(
        fecha_desde=datetime.fromisoformat(filtros['fecha_desde']).date() if filtros.get('fecha_desde') else None,
        fecha_hasta=datetime.fromisoformat(filtros['fecha_hasta']).date() if filtros.get('fecha_hasta') else None,
        cliente_filtro=cliente_filtro,
        razon_social_filtro=filtros.get('razon_social'),
        estado_venta_filtro=filtros.get('estado_venta'),
        tipo_factura_filtro=filtros.get('tipo_factura'),
    )
    
    # Calcular totales
    total_blanco = sum(float(i['monto']) for i in ventas_reporte if i['tipo'] == 'Blanco')
    total_negro = sum(float(i['monto']) for i in ventas_reporte if i['tipo'] == 'Negro')
    total = total_blanco + total_negro
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Ventas"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    
    # T�tulo
    ws['A1'] = 'REPORTE DE VENTAS'
    ws['A1'].font = title_font
    ws.merge_cells('A1:H1')
    
    # Resumen
    ws['A3'] = 'Total Ventas Blanco:'
    ws['B3'] = total_blanco
    ws['B3'].number_format = '$#,##0.00'
    ws['A3'].font = Font(bold=True)
    
    ws['A4'] = 'Total Ventas Negro:'
    ws['B4'] = total_negro
    ws['B4'].number_format = '$#,##0.00'
    ws['A4'].font = Font(bold=True)
    
    ws['A5'] = 'TOTAL VENTAS:'
    ws['B5'] = total
    ws['B5'].number_format = '$#,##0.00'
    ws['A5'].font = Font(bold=True, size=12)
    ws['B5'].font = Font(bold=True, size=12)
    
    # Headers
    headers = ['Fecha Venta', 'Pedido', 'ID Factura', 'Cliente', 'Raz�n Social', 'Forma Pago', 'Monto', 'Tipo']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(7, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    for row, ingreso in enumerate(ventas_reporte, 8):
        ws.cell(row, 1, ingreso['fecha'].strftime('%d/%m/%Y'))
        ws.cell(row, 2, ingreso['pedido'])
        ws.cell(row, 3, ingreso['numero_factura'])
        ws.cell(row, 4, ingreso['cliente'])
        ws.cell(row, 5, ingreso['razon_social'])
        ws.cell(row, 6, ingreso['forma_pago'])
        cell = ws.cell(row, 7, ingreso['monto'])
        cell.number_format = '$#,##0.00'
        ws.cell(row, 8, ingreso['tipo'])
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 10
    
    # Respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reporte_ventas.xlsx'
    wb.save(response)
    return response


@login_required
def reportes_gastos(request):
    from datetime import datetime
    from .forms import ReporteGastosForm
    
    form = ReporteGastosForm()
    
    fecha_desde = None
    fecha_hasta = None
    cuenta_filtro = None
    tipo_cuenta_filtro = None
    tipo_factura_filtro = None
    
    if request.method == 'POST':
        form = ReporteGastosForm(request.POST)
        if form.is_valid():
            fecha_desde = form.cleaned_data.get('fecha_desde')
            fecha_hasta = form.cleaned_data.get('fecha_hasta')
            cuenta_filtro = form.cleaned_data.get('cuenta')
            tipo_cuenta_filtro = form.cleaned_data.get('tipo_cuenta')
            tipo_factura_filtro = form.cleaned_data.get('tipo_factura')
            
            request.session['reporte_gastos_filtros'] = {
                'fecha_desde': fecha_desde.isoformat() if fecha_desde else None,
                'fecha_hasta': fecha_hasta.isoformat() if fecha_hasta else None,
                'cuenta_id': [c.id for c in cuenta_filtro] if cuenta_filtro else None,
                'tipo_cuenta_id': [t.id for t in tipo_cuenta_filtro] if tipo_cuenta_filtro else None,
                'tipo_factura': list(tipo_factura_filtro) if tipo_factura_filtro else None,
            }
    
    gastos = []
    compras_query = Compra.objects.filter(deleted_at__isnull=True).select_related('cuenta', 'cuenta__tipo_cuenta')
    
    if fecha_desde:
        compras_query = compras_query.filter(fecha_pago__gte=fecha_desde)
    if fecha_hasta:
        compras_query = compras_query.filter(fecha_pago__lte=fecha_hasta)
    if cuenta_filtro:
        compras_query = compras_query.filter(cuenta__in=cuenta_filtro)
    if tipo_cuenta_filtro:
        compras_query = compras_query.filter(cuenta__tipo_cuenta__in=tipo_cuenta_filtro)
    if tipo_factura_filtro:
        if 'blanco' in tipo_factura_filtro and 'negro' not in tipo_factura_filtro:
            compras_query = compras_query.filter(con_factura=True)
        elif 'negro' in tipo_factura_filtro and 'blanco' not in tipo_factura_filtro:
            compras_query = compras_query.filter(con_factura=False)
    
    for compra in compras_query:
        gastos.append({
            'fecha': compra.fecha_pago,
            'numero_pedido': compra.numero_pedido or '-',
            'numero_factura': compra.numero_factura or '-',
            'cuenta': str(compra.cuenta),
            'tipo_cuenta': compra.cuenta.tipo_cuenta.get_tipo_display(),
            'monto': compra.valor_total,
            'tipo': 'Blanco' if compra.con_factura else 'Negro'
        })
    
    gastos.sort(key=lambda x: x['fecha'], reverse=True)
    
    total_blanco = sum(g['monto'] for g in gastos if g['tipo'] == 'Blanco')
    total_negro = sum(g['monto'] for g in gastos if g['tipo'] == 'Negro')
    total = total_blanco + total_negro
    
    cantidad_blanco = len([g for g in gastos if g['tipo'] == 'Blanco'])
    cantidad_negro = len([g for g in gastos if g['tipo'] == 'Negro'])
    
    reporte_data = {
        'gastos': {
            'total_blanco': total_blanco,
            'total_negro': total_negro,
            'total': total,
            'cantidad_blanco': cantidad_blanco,
            'cantidad_negro': cantidad_negro,
            'cantidad': len(gastos),
            'lista': gastos
        }
    }
    
    context = {
        'form': form,
        'reporte_data': reporte_data,
    }
    return render(request, 'comercial/reportes/reportes_gastos.html', context)


@login_required
def reportes_proveedores(request):
    form = ReporteProveedorForm(request.GET or None)
    proveedores = Cuenta.objects.filter(
        deleted_at__isnull=True,
        activo=True,
        tipo_cuenta__tipo='proveedores',
    ).select_related('tipo_cuenta').order_by('nombre')

    proveedor_filtro = None
    if form.is_valid() and form.cleaned_data.get('proveedor'):
        proveedor_filtro = form.cleaned_data['proveedor']
        proveedores = proveedores.filter(pk=proveedor_filtro.pk)

    resumen_proveedores = []
    for proveedor in proveedores:
        cuenta_corriente = construir_cuenta_corriente_proveedor(proveedor)
        resumen_proveedores.append(cuenta_corriente)

    context = {
        'form': form,
        'resumen_proveedores': resumen_proveedores,
        'proveedor_filtro': proveedor_filtro,
    }
    return render(request, 'comercial/reportes/reportes_proveedores.html', context)


@login_required
def reporte_proveedor_detalle(request, pk):
    proveedor = get_object_or_404(
        Cuenta.objects.filter(
            deleted_at__isnull=True,
            activo=True,
            tipo_cuenta__tipo='proveedores',
        ).select_related('tipo_cuenta'),
        pk=pk,
    )
    cuenta_corriente = construir_cuenta_corriente_proveedor(proveedor)

    context = {
        'cuenta_corriente': cuenta_corriente,
    }
    return render(request, 'comercial/reportes/reporte_proveedor_detalle.html', context)


@login_required
def exportar_reporte_proveedores_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse

    proveedor_id = request.GET.get('proveedor')
    proveedor = get_object_or_404(
        Cuenta.objects.filter(
            deleted_at__isnull=True,
            activo=True,
            tipo_cuenta__tipo='proveedores',
        ).select_related('tipo_cuenta'),
        pk=proveedor_id,
    )

    cuenta_corriente = construir_cuenta_corriente_proveedor(proveedor)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Cuenta Proveedor'

    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=11)
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')

    ws['A1'] = 'CUENTA CORRIENTE DE PROVEEDOR'
    ws['A1'].font = title_font
    ws.merge_cells('A1:G1')

    ws['A3'] = 'Proveedor:'
    ws['B3'] = proveedor.nombre
    ws['A4'] = 'Razón Social:'
    ws['B4'] = proveedor.razon_social or '-'
    ws['A5'] = 'Total Compras:'
    ws['B5'] = float(cuenta_corriente['total_compras'])
    ws['A6'] = 'Total Señas:'
    ws['B6'] = float(cuenta_corriente['total_senas'])
    ws['A7'] = 'Total Pagos:'
    ws['B7'] = float(cuenta_corriente['total_pagos'])
    ws['A8'] = 'Saldo Actual:'
    ws['B8'] = float(cuenta_corriente['saldo_actual'])

    for cell_ref in ('A3', 'A4', 'A5', 'A6', 'A7', 'A8'):
        ws[cell_ref].font = section_font

    for cell_ref in ('B5', 'B6', 'B7', 'B8'):
        ws[cell_ref].number_format = '$#,##0.00'

    headers = ['Fecha', 'Movimiento', 'Referencia', 'Forma de Pago', 'Debe', 'Haber', 'Saldo']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(10, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row, movimiento in enumerate(cuenta_corriente['movimientos'], 11):
        ws.cell(row, 1, movimiento['fecha'].strftime('%d/%m/%Y'))
        ws.cell(row, 2, movimiento['descripcion'])
        ws.cell(row, 3, movimiento['referencia'])
        ws.cell(row, 4, movimiento['forma_pago'])

        debe_cell = ws.cell(row, 5, float(movimiento['debe']))
        haber_cell = ws.cell(row, 6, float(movimiento['haber']))
        saldo_cell = ws.cell(row, 7, float(movimiento['saldo']))
        for money_cell in (debe_cell, haber_cell, saldo_cell):
            money_cell.number_format = '$#,##0.00'

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 34
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=cuenta_proveedor_{proveedor.id}.xlsx'
    wb.save(response)
    return response


@login_required
def exportar_reporte_gastos_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse
    
    gastos = []
    compras_query = Compra.objects.filter(deleted_at__isnull=True).select_related('cuenta', 'cuenta__tipo_cuenta')
    
    for compra in compras_query:
        gastos.append({
            'fecha': compra.fecha_pago,
            'numero_pedido': compra.numero_pedido or '-',
            'numero_factura': compra.numero_factura or '-',
            'cuenta': str(compra.cuenta),
            'tipo_cuenta': compra.cuenta.tipo_cuenta.get_tipo_display(),
            'monto': float(compra.valor_total),
            'tipo': 'Blanco' if compra.con_factura else 'Negro'
        })
    
    gastos.sort(key=lambda x: x['fecha'], reverse=True)
    
    total_blanco = sum(g['monto'] for g in gastos if g['tipo'] == 'Blanco')
    total_negro = sum(g['monto'] for g in gastos if g['tipo'] == 'Negro')
    total = total_blanco + total_negro
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Gastos"
    
    header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    
    ws['A1'] = 'REPORTE DE GASTOS'
    ws['A1'].font = title_font
    ws.merge_cells('A1:G1')
    
    ws['A3'] = 'Total Gastos Blanco:'
    ws['B3'] = total_blanco
    ws['B3'].number_format = '$#,##0.00'
    ws['A3'].font = Font(bold=True)
    
    ws['A4'] = 'Total Gastos Negro:'
    ws['B4'] = total_negro
    ws['B4'].number_format = '$#,##0.00'
    ws['A4'].font = Font(bold=True)
    
    ws['A5'] = 'TOTAL GASTOS:'
    ws['B5'] = total
    ws['B5'].number_format = '$#,##0.00'
    ws['A5'].font = Font(bold=True, size=12)
    ws['B5'].font = Font(bold=True, size=12)
    
    headers = ['Fecha Pago', 'N� Pedido', 'N� Factura', 'Cuenta', 'Tipo Cuenta', 'Monto', 'Tipo']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(7, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    for row, gasto in enumerate(gastos, 8):
        ws.cell(row, 1, gasto['fecha'].strftime('%d/%m/%Y'))
        ws.cell(row, 2, gasto['numero_pedido'])
        ws.cell(row, 3, gasto['numero_factura'])
        ws.cell(row, 4, gasto['cuenta'])
        ws.cell(row, 5, gasto['tipo_cuenta'])
        cell = ws.cell(row, 6, gasto['monto'])
        cell.number_format = '$#,##0.00'
        ws.cell(row, 7, gasto['tipo'])
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 10
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reporte_gastos.xlsx'
    wb.save(response)
    return response


@login_required
def editar_fecha_sena(request, pk):
    import json
    from datetime import datetime
    
    if request.method == 'POST':
        venta = get_object_or_404(Venta, pk=pk)
        
        try:
            if request.POST:
                fecha_sena_str = request.POST.get('fecha') or request.POST.get('fecha_sena')
            else:
                data = json.loads(request.body)
                fecha_sena_str = data.get('fecha') or data.get('fecha_sena')

            if not fecha_sena_str:
                return JsonResponse({'success': False, 'error': 'Debe indicar una fecha valida'}, status=400)
            
            # Convertir fecha string a objeto datetime
            if isinstance(fecha_sena_str, str):
                fecha_sena = datetime.strptime(fecha_sena_str, '%Y-%m-%d')
            else:
                return JsonResponse({'success': False, 'error': 'Formato de fecha invalido'}, status=400)
            
            # Actualizar created_at (fecha de se�a) y fecha_pago (fecha de venta total)
            venta.created_at = fecha_sena
            venta.fecha_pago = fecha_sena.date()
            venta.save()
            
            return JsonResponse({
                'success': True,
                'fecha': fecha_sena.strftime('%d/%m/%Y')
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    
    return JsonResponse({'success': False, 'error': 'Metodo no permitido'}, status=405)


@login_required
def eliminar_pago(request, pk):
    import json
    
    if request.method == 'POST':
        pago = get_object_or_404(PagoVenta, pk=pk)
        venta = pago.venta
        
        try:
            # Eliminar el pago
            pago.delete()
            
            # Recalcular saldo de la venta
            venta.save()
            
            return JsonResponse({
                'success': True,
                'saldo': float(venta.saldo)
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    
    return JsonResponse({'success': False, 'error': 'Metodo no permitido'}, status=405)


@login_required
def agregar_retencion_pago(request, pk):
    """Agregar retenci�n a un pago existente"""
    from .models import Retencion
    import json
    
    if request.method == 'POST':
        pago = get_object_or_404(PagoVenta, pk=pk)
        
        try:
            data = json.loads(request.body)
            tipo = data.get('tipo')
            importe = Decimal(str(data.get('importe')))
            concepto = data.get('concepto', '')
            numero_comprobante = data.get('numero_comprobante', '')
            
            if importe <= 0:
                return JsonResponse({'success': False, 'error': 'El importe debe ser mayor a 0'}, status=400)
            
            retencion = Retencion.objects.create(
                pago=pago,
                tipo=tipo,
                concepto=concepto,
                numero_comprobante=numero_comprobante,
                importe_retenido=importe,
                fecha_comprobante=pago.fecha_pago
            )
            
            return JsonResponse({
                'success': True,
                'retencion': {
                    'id': retencion.id,
                    'tipo': retencion.get_tipo_display(),
                    'importe': float(retencion.importe_retenido)
                }
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    
    return JsonResponse({'success': False, 'error': 'Metodo no permitido'}, status=405)






@login_required
def reporte_general(request):
    """Vista para reporte general combinando ingresos y gastos."""
    from django.db.models import Sum, Value, DecimalField, Q
    from django.db.models.functions import Coalesce
    
    reporte_data = None
    
    if request.method == 'POST':
        fecha_desde = request.POST.get('fecha_desde')
        fecha_hasta = request.POST.get('fecha_hasta')
        
        # INGRESOS = Señas + Pagos adicionales
        # 1. Señas de ventas
        ventas_senas = Venta.objects.filter(deleted_at__isnull=True, sena__gt=0)
        if fecha_desde:
            ventas_senas = ventas_senas.filter(created_at__date__gte=fecha_desde)
        if fecha_hasta:
            ventas_senas = ventas_senas.filter(created_at__date__lte=fecha_hasta)
        
        senas_blanco = ventas_senas.filter(con_factura=True).aggregate(
            total=Coalesce(Sum('sena'), Value(0, output_field=DecimalField()))
        )['total']
        senas_negro = ventas_senas.filter(con_factura=False).aggregate(
            total=Coalesce(Sum('sena'), Value(0, output_field=DecimalField()))
        )['total']
        
        # 2. Pagos adicionales
        pagos = PagoVenta.objects.filter(venta__deleted_at__isnull=True)
        if fecha_desde:
            pagos = pagos.filter(fecha_pago__gte=fecha_desde)
        if fecha_hasta:
            pagos = pagos.filter(fecha_pago__lte=fecha_hasta)
        
        pagos_blanco = pagos.filter(con_factura=True).aggregate(
            total=Coalesce(Sum('monto'), Value(0, output_field=DecimalField()))
        )['total']
        pagos_negro = pagos.filter(con_factura=False).aggregate(
            total=Coalesce(Sum('monto'), Value(0, output_field=DecimalField()))
        )['total']
        
        # Total ingresos = señas + pagos
        ingresos_blanco = senas_blanco + pagos_blanco
        ingresos_negro = senas_negro + pagos_negro
        total_ingresos = ingresos_blanco + ingresos_negro
        
        # GASTOS (Compra)
        compras = Compra.objects.filter(deleted_at__isnull=True)
        if fecha_desde:
            compras = compras.filter(fecha_pago__gte=fecha_desde)
        if fecha_hasta:
            compras = compras.filter(fecha_pago__lte=fecha_hasta)
        
        gastos_blanco = compras.filter(con_factura=True).aggregate(
            total=Coalesce(Sum('valor_total'), Value(0, output_field=DecimalField()))
        )['total']
        gastos_negro = compras.filter(con_factura=False).aggregate(
            total=Coalesce(Sum('valor_total'), Value(0, output_field=DecimalField()))
        )['total']
        total_gastos = gastos_blanco + gastos_negro
        
        # Balance
        balance_blanco = ingresos_blanco - gastos_blanco
        balance_negro = ingresos_negro - gastos_negro
        balance_total = total_ingresos - total_gastos
        
        reporte_data = {
            'ingresos': {
                'blanco': ingresos_blanco,
                'negro': ingresos_negro,
                'total': total_ingresos,
                'cantidad': ventas_senas.count() + pagos.count(),
            },
            'gastos': {
                'blanco': gastos_blanco,
                'negro': gastos_negro,
                'total': total_gastos,
                'cantidad': compras.count(),
            },
            'balance': {
                'blanco': balance_blanco,
                'negro': balance_negro,
                'total': balance_total,
            },
        }
        
        request.session['reporte_general_data'] = {
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
            'ingresos_blanco': float(ingresos_blanco),
            'ingresos_negro': float(ingresos_negro),
            'total_ingresos': float(total_ingresos),
            'gastos_blanco': float(gastos_blanco),
            'gastos_negro': float(gastos_negro),
            'total_gastos': float(total_gastos),
            'balance_blanco': float(balance_blanco),
            'balance_negro': float(balance_negro),
            'balance_total': float(balance_total),
        }
    
    return render(request, 'comercial/reportes/reporte_general.html', {'reporte_data': reporte_data})


@login_required
def exportar_reporte_general_excel(request):
    """Exportar reporte general a Excel."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from django.http import HttpResponse
    from datetime import datetime
    
    data = request.session.get('reporte_general_data', {})
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte General"
    
    # Estilos
    header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # T�tulo
    ws.merge_cells('A1:E1')
    ws['A1'] = 'REPORTE GENERAL - INGRESOS Y GASTOS'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Per�odo
    row = 3
    if data.get('fecha_desde') or data.get('fecha_hasta'):
        periodo = f"Per�odo: {data.get('fecha_desde', 'Inicio')} - {data.get('fecha_hasta', 'Hoy')}"
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = periodo
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 2
    
    # Encabezados
    headers = ['Concepto', 'Blanco', 'Negro', 'Total', 'Porcentaje']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    row += 1
    
    # INGRESOS
    ws.cell(row=row, column=1, value='INGRESOS').font = Font(bold=True)
    ws.cell(row=row, column=2, value=data.get('ingresos_blanco', 0))
    ws.cell(row=row, column=3, value=data.get('ingresos_negro', 0))
    ws.cell(row=row, column=4, value=data.get('total_ingresos', 0))
    ws.cell(row=row, column=5, value='100%')
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = border
        if col > 1:
            ws.cell(row=row, column=col).number_format = '"$"#,##0.00'
    row += 1
    
    # GASTOS
    ws.cell(row=row, column=1, value='GASTOS').font = Font(bold=True)
    ws.cell(row=row, column=2, value=data.get('gastos_blanco', 0))
    ws.cell(row=row, column=3, value=data.get('gastos_negro', 0))
    ws.cell(row=row, column=4, value=data.get('total_gastos', 0))
    total_ingresos = data.get('total_ingresos', 0)
    porcentaje_gastos = (data.get('total_gastos', 0) / total_ingresos * 100) if total_ingresos > 0 else 0
    ws.cell(row=row, column=5, value=f'{porcentaje_gastos:.1f}%')
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = border
        if col > 1 and col < 5:
            ws.cell(row=row, column=col).number_format = '"$"#,##0.00'
    row += 1
    
    # BALANCE
    balance_fill = PatternFill(start_color="D4EDDA" if data.get('balance_total', 0) >= 0 else "F8D7DA", 
                               end_color="D4EDDA" if data.get('balance_total', 0) >= 0 else "F8D7DA", 
                               fill_type="solid")
    ws.cell(row=row, column=1, value='BALANCE').font = Font(bold=True, size=12)
    ws.cell(row=row, column=2, value=data.get('balance_blanco', 0))
    ws.cell(row=row, column=3, value=data.get('balance_negro', 0))
    ws.cell(row=row, column=4, value=data.get('balance_total', 0))
    ws.cell(row=row, column=5, value='')
    for col in range(1, 6):
        ws.cell(row=row, column=col).fill = balance_fill
        ws.cell(row=row, column=col).border = border
        ws.cell(row=row, column=col).font = Font(bold=True)
        if col > 1 and col < 5:
            ws.cell(row=row, column=col).number_format = '"$"#,##0.00'
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'reporte_general_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


